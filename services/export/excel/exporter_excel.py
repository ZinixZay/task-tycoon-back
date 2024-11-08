from typing import List, Optional
from uuid import UUID
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from services.export.excel.helper_excel import rows, columns 
from models import AttemptStatsModel, QuestionModel, TaskModel
from repositories import AttemptStatsRepository, QuestionRepository, UserRepository
from utils.enums.attempt_type_enum import AttemptStatsStatusEnum
from utils.helpers import map_by


class ExporterExcel:
    task: TaskModel
    questions: List[QuestionModel]
    user_ids: Optional[List[UUID]]
    attempt_ids: Optional[List[UUID]]
    wb: Workbook
    c: str
    r: str
    cell: str
    
    resulting_attempts: Optional[List[AttemptStatsModel]]
    
    default_side: Side
    default_border: Border
    default_alignment: Alignment
    header_font: Font
    header2_font: Font
    label_font: Font
    header_fill: PatternFill
    header2_fill: PatternFill
    header3_fill: PatternFill
    t_answer_fill: PatternFill
    f_answer_fill: PatternFill
    n_answer_fill: PatternFill
    
    
    
    @classmethod
    async def create(cls, task: TaskModel, user_ids: List[UUID] = None, attempt_ids: List[UUID] = None):
        exporter_excel = ExporterExcel()
        exporter_excel.task = task
        exporter_excel.user_ids = user_ids
        exporter_excel.attempt_ids = attempt_ids
        exporter_excel.questions = await QuestionRepository.find_by_task(task.id) 
        
        exporter_excel.default_side = Side(border_style='thin', color='FF000000')
        exporter_excel.default_border = Border(left=exporter_excel.default_side, 
                                               right=exporter_excel.default_side,
                                               top=exporter_excel.default_side,
                                               bottom=exporter_excel.default_side)
        exporter_excel.default_alignment = Alignment(horizontal='center', vertical='center') 
        exporter_excel.header_font = Font(bold=True, size=14, name='Arial')
        exporter_excel.header2_font = Font(bold=True, size=11, name='Arial')
        exporter_excel.label_font = Font(size=11, name='Calibri')
        exporter_excel.header_fill = PatternFill(start_color='ffe699', end_color='ffe699', fill_type='solid')
        exporter_excel.header2_fill = PatternFill(start_color='8ea9db', end_color='8ea9db', fill_type='solid')
        exporter_excel.header3_fill = PatternFill(start_color='b4c6e7', end_color='b4c6e7', fill_type='solid')
        exporter_excel.t_answer_fill = PatternFill(start_color='c6e0b4', end_color='c6e0b4', fill_type='solid')
        exporter_excel.f_answer_fill = PatternFill(start_color='ff9b9b', end_color='ff9b9b', fill_type='solid')
        exporter_excel.n_answer_fill = PatternFill(start_color='bfbfbf', end_color='bfbfbf', fill_type='solid')
        
        return exporter_excel
    
    async def export(self):
        self.wb = Workbook()
        answers_sheet = self.wb.active
        answers_sheet.title = 'Ответы'
        
        await self.__draw_answer_sheet__()
        
        active_ws = self.wb.active
        self.wb.remove(active_ws)
        self.wb.save(f"resources/excel/task_{self.task.id}.xlsx")
    
    async def __draw_answer_sheet__(self):
        # sheet
        self.wb.create_sheet("Ответы", 0)
        answer_sheet: Worksheet = self.wb['Ответы']
        self.__reset_cell__()
        
        # title
        answer_sheet.merge_cells(f'{self.cell}:{self.__column_plus__(len(self.questions) + 1, False)+self.r}')
        self.__stylize_cell__(answer_sheet[self.cell], font=self.header_font, fill=self.header_fill)
        answer_sheet[self.cell] = 'Вопросы'

        # header 
        self.__row_down__()
        self.__stylize_cell__(answer_sheet[self.cell], font=self.header2_font, fill=self.header2_fill)
        answer_sheet[self.cell] = 'Пользователи'
        for question_id in range(1, len(self.questions) + 1):
            self.__column_plus__()
            answer_sheet[self.cell] = question_id
            self.__stylize_cell__(answer_sheet[self.cell], font=self.label_font, fill=self.header2_fill)
        self.__column_plus__()
        answer_sheet[self.cell] = "Верно"
        self.__stylize_cell__(answer_sheet[self.cell], font=self.header2_font, fill=self.header2_fill)
        
        # results
        self.__row_down__()
        resulting_attempts_stats = await AttemptStatsRepository.find_resulting_by_task(self.task.id)
        user_entities = await UserRepository.find(list(map(lambda x: x.user_id, resulting_attempts_stats)))
        users_map = map_by(user_entities, lambda x: x.id)
        for stats in resulting_attempts_stats:
            user = users_map[stats.user_id]
            answer_sheet[self.cell] = f'{user.name} {user.surname}'
            self.__stylize_cell__(answer_sheet[self.cell], font=self.header2_font, fill=self.header3_fill)
            c_correct = 0
            for question_stats in stats.stats:
                self.__column_plus__()
                if question_stats['status'] == AttemptStatsStatusEnum.correct.value:
                    c_correct += 1
                    self.__stylize_cell__(answer_sheet[self.cell], fill=self.t_answer_fill)
                elif question_stats['status'] == AttemptStatsStatusEnum.no_answer.value:
                    self.__stylize_cell__(answer_sheet[self.cell], fill=self.n_answer_fill)
                elif question_stats['status'] == AttemptStatsStatusEnum.wrong.value:
                    self.__stylize_cell__(answer_sheet[self.cell], fill=self.f_answer_fill)
            self.__column_plus__()
            answer_sheet[self.cell] = f'{c_correct}'
            self.__stylize_cell__(answer_sheet[self.cell], font=self.header2_font)
            self.__row_down__()
        
        
        # formatting
        self.__set_column_width_automatically__(answer_sheet)
    
    def __set_column_width_automatically__(self, sheet: Worksheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = column[1].column_letter
            for cell in column:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 5)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
    def __stylize_cell__(
        self, 
        cell: Cell, 
        alignment: Alignment = None, 
        font: Font = None, 
        fill: PatternFill = None, 
        border: Border = None
        ):
        cell.font = font if font else self.label_font
        cell.fill = fill if fill else self.header2_fill
        cell.border = border if border else self.default_border
        cell.alignment = alignment if alignment else self.default_alignment
    
    def __reset_row__(self):
        self.r = rows[0]
        
    def __reset_column__(self):
        self.c = columns[0]
        
    def __reset_cell__(self):
        self.c, self.r = columns[0], rows[0]
        self.cell = self.c + self.r
        
    def __column_right__(self, value: int = 1) -> str:
        self.__reset_row__()
        self.c = self.__column_minus__(value)
        self.__set_cell__()    
    
    def __column_right__(self, value: int = 1) -> str:
        self.__reset_row__()
        self.c = self.__column_plus__(value)
        self.__set_cell__()
        
    def __row_down__(self, value: int = 1) -> str:
        self.__reset_column__()
        self.r = self.__row_plus__(value)
        self.__set_cell__()
    
    def __row_up__(self, value: int = 1) -> str:
        self.__reset_column__()
        self.r = self.__row_minus__(value)
        self.__set_cell__()
    
    def __column_plus__(self, value: int = 1, sync: bool = True) -> str:
        new_column = columns[columns.index(self.c) + value]
        if sync:
            self.c = new_column
            self.__set_cell__()
        return new_column
    
    def __column_minus__(self, value: int = 1, sync: bool = True) -> str:
        new_column = columns[columns.index(self.c) - value]
        if sync:
            self.c = new_column
            self.__set_cell__()
        return new_column
    
    def __row_plus__(self, value: int = 1, sync: bool = True) -> str:
        new_row = rows[rows.index(self.r) + value]
        if sync:
            self.r = new_row
            self.__set_cell__()
        return new_row
        
    def __row_minus__(self, value: int = 1, sync: bool = True) -> str:
        new_row = rows[rows.index(self.r) - value]
        if sync:
            self.r = new_row
            self.__set_cell__()
        return new_row

    def __set_cell__(self):
        self.cell = self.c + self.r
    